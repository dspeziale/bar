# -*- coding: utf-8 -*-
"""Caricamento massivo da Excel: modelli .xlsx da scaricare, compilare e
ricaricare. Un modello per tipo (prodotti, ingredienti del builder, articoli
del banco, consumabili del magazzino, clienti, piatti delle convenzioni).

Ogni modello ha tre fogli: `Dati` (solo l'intestazione, da riempire),
`Esempio` (due righe compilate, che non vengono importate) e `Istruzioni`
(colonna per colonna: obbligatoria o no, che valori accetta, le categorie e
gli allergeni esistenti). L'importazione riconosce le colonne dal nome, non
dalla posizione, quindi il gestore può spostarle o togliere quelle che non
usa. Una riga con lo stesso nome di una già presente la aggiorna; le righe
con errori vengono saltate e spiegate una per una.
"""
import io
import re
import secrets
import unicodedata
from datetime import datetime, date

from app import db
from app.models import (Product, Category, Ingredient, IngredientCategory, BancoItem,
                        ConsumableItem, Supplier, User, CorporateAccount,
                        CorporateMembership, MealConfiguration, ALLERGENS)

SI = ('si', 'sì', 's', '1', 'x', 'true', 'vero', 'yes', 'y')
NO = ('no', 'n', '0', '', 'false', 'falso')

# Colonne: (intestazione, obbligatoria, spiegazione, tipo)
# tipo: testo / numero / intero / sino / allergeni / data / elenco:a,b,c
MODELLI = [
    ('prodotti', 'Prodotti del listino', 'fa-burger',
     'Piatti, panini, bevande e tutto quello che il cliente ordina dal menu. La categoria che '
     'non esiste viene creata.',
     [('Nome', True, 'Il nome come compare nel menu. Se esiste già, la riga aggiorna il prodotto.', 'testo'),
      ('Categoria', True, 'Una categoria del listino; se non esiste viene creata.', 'testo'),
      ('Prezzo', True, 'In euro, con virgola o punto (es. 4,50).', 'numero'),
      ('Descrizione', False, 'Testo libero mostrato sotto il nome.', 'testo'),
      ('Allergeni', False, 'Chiavi separate da virgola (vedi elenco): glutine, latte, uova…', 'allergeni'),
      ('Quantita giornaliera', False, 'Pezzi disponibili al giorno (default 20).', 'intero'),
      ('Vegetariano', False, 'SI / NO', 'sino'),
      ('Vegano', False, 'SI / NO', 'sino'),
      ('Kcal', False, 'Calorie per porzione (vuoto = non indicato).', 'intero'),
      ('Proteine g', False, 'Grammi per porzione.', 'numero'),
      ('Carboidrati g', False, 'Grammi per porzione.', 'numero'),
      ('Grassi g', False, 'Grammi per porzione.', 'numero'),
      ('Codice a barre', False, 'EAN o codice interno, facoltativo.', 'testo'),
      ('Attivo', False, 'SI / NO (default SI).', 'sino')],
     [['Insalata di farro', 'Insalate', '6,50', 'Farro, pomodorini, feta, olive', 'glutine, latte',
       '15', 'SI', 'NO', '420', '14', '58', '14', '', 'SI'],
      ['Acqua naturale 50cl', 'Bevande', '1,00', '', '', '60', 'SI', 'SI', '0', '0', '0', '0',
       '8001234567890', 'SI']]),
    ('ingredienti', 'Ingredienti del builder', 'fa-seedling',
     'Gli ingredienti con cui il cliente compone panino, insalata e poke. La categoria si '
     'riconosce da nome e tipo di builder.',
     [('Nome', True, 'Il nome dell\'ingrediente.', 'testo'),
      ('Categoria', True, 'Es. Pane, Proteine, Verdure, Salse. Se non esiste viene creata.', 'testo'),
      ('Builder', True, 'panino, insalata o poke.', 'elenco:panino,insalata,poke'),
      ('Prezzo extra', False, 'Supplemento in euro (default 0).', 'numero'),
      ('Allergeni', False, 'Chiavi separate da virgola.', 'allergeni'),
      ('Vegetariano', False, 'SI / NO', 'sino'),
      ('Vegano', False, 'SI / NO', 'sino'),
      ('Kcal', False, 'Calorie per porzione.', 'intero'),
      ('Proteine g', False, '', 'numero'),
      ('Carboidrati g', False, '', 'numero'),
      ('Grassi g', False, '', 'numero'),
      ('Grammi per porzione', False, 'Per il magazzino, facoltativo.', 'numero'),
      ('Attivo', False, 'SI / NO (default SI).', 'sino')],
     [['Pane integrale', 'Pane', 'panino', '0', 'glutine', 'SI', 'SI', '180', '7', '34', '2', '90', 'SI'],
      ['Salmone', 'Proteine', 'poke', '1,50', 'pesce', 'NO', 'NO', '120', '18', '0', '6', '60', 'SI']]),
    ('banco', 'Articoli del banco', 'fa-mug-hot',
     'Caffè, cornetti e consumazioni veloci battute al banco (POS).',
     [('Nome', True, 'Il nome sul tasto del banco.', 'testo'),
      ('Prezzo', True, 'In euro.', 'numero'),
      ('Icona', False, 'Classe Font Awesome, es. fa-mug-hot, fa-cookie (default fa-mug-hot).', 'testo'),
      ('Colore', False, 'primary, secondary, success, danger, warning, info, dark (default info).',
       'elenco:primary,secondary,success,danger,warning,info,dark'),
      ('Ordine', False, 'Numero per l\'ordinamento dei tasti.', 'intero'),
      ('Attivo', False, 'SI / NO (default SI).', 'sino')],
     [['Caffè', '1,20', 'fa-mug-hot', 'dark', '1', 'SI'],
      ['Cornetto', '1,30', 'fa-cookie', 'warning', '2', 'SI']]),
    ('consumabili', 'Consumabili del magazzino', 'fa-warehouse',
     'Bicchieri, tovaglioli, caffè in grani: le scorte con soglia di riordino e fornitore.',
     [('Nome', True, 'L\'articolo.', 'testo'),
      ('Unita', False, 'pz, kg, lt, conf… (default pz).', 'testo'),
      ('Quantita', False, 'Giacenza attuale.', 'numero'),
      ('Soglia minima', False, 'Sotto questa quantità scatta l\'avviso (0 = nessun avviso).', 'numero'),
      ('Fornitore', False, 'Nome del fornitore; se non esiste viene creato.', 'testo')],
     [['Bicchieri caffè', 'pz', '1200', '300', 'Cartotecnica Rossi'],
      ['Caffè in grani', 'kg', '6', '2', 'Torrefazione Bianchi']]),
    ('clienti', 'Clienti', 'fa-users',
     'L\'anagrafica dei clienti, ad esempio i dipendenti di un\'azienda convenzionata. A ogni '
     'cliente nuovo viene assegnata una password provvisoria, riportata nel resoconto.',
     [('Nome', True, '', 'testo'),
      ('Cognome', True, '', 'testo'),
      ('Email', True, 'È anche il nome utente per l\'accesso; se esiste già, aggiorna la scheda.', 'testo'),
      ('Telefono', False, '', 'testo'),
      ('Data di nascita', False, 'GG/MM/AAAA (serve per gli auguri e per la dieta).', 'data'),
      ('Reparto', False, 'Ufficio o piano, per consegne e ritiri.', 'testo'),
      ('Azienda convenzionata', False, 'Nome esatto di una convenzione esistente: il cliente vi viene iscritto.', 'testo'),
      ('Attivo', False, 'SI / NO (default SI: il cliente può accedere subito).', 'sino')],
     [['Maria', 'Rossi', 'maria.rossi@esempio.it', '+39 333 1234567', '12/05/1988', 'Amministrazione', '', 'SI'],
      ['Luca', 'Bianchi', 'luca.bianchi@esempio.it', '', '', 'Magazzino, piano -1', 'Acme S.p.A.', 'SI']]),
    ('pasti_convenzione', 'Piatti delle convenzioni', 'fa-building',
     'Le configurazioni di pasto (primo, secondo, contorno…) riutilizzabili per il pasto del giorno '
     'di una convenzione aziendale. L\'azienda deve esistere già.',
     [('Azienda', True, 'Nome esatto della convenzione.', 'testo'),
      ('Nome', True, 'Il nome del pasto (es. Menu completo, Menu leggero).', 'testo'),
      ('Primo', False, '', 'testo'),
      ('Secondo', False, '', 'testo'),
      ('Contorno', False, '', 'testo'),
      ('Bevanda', False, '', 'testo'),
      ('Caffe', False, '', 'testo'),
      ('Descrizione', False, '', 'testo'),
      ('Allergeni', False, 'Chiavi separate da virgola.', 'allergeni'),
      ('Prezzo', False, 'Vuoto = prezzo della convenzione.', 'numero'),
      ('Max prenotazioni', False, 'Vuoto = coperti della convenzione.', 'intero'),
      ('Vegetariano', False, 'SI / NO', 'sino'),
      ('Vegano', False, 'SI / NO', 'sino'),
      ('Kcal', False, '', 'intero'),
      ('Proteine g', False, '', 'numero'),
      ('Carboidrati g', False, '', 'numero'),
      ('Grassi g', False, '', 'numero')],
     [['Acme S.p.A.', 'Menu completo', 'Pasta al pomodoro', 'Pollo alla griglia', 'Verdure grigliate',
       'Acqua', 'Caffè', '', 'glutine', '', '', 'NO', 'NO', '780', '38', '92', '22'],
      ['Acme S.p.A.', 'Menu leggero', 'Insalata di farro', '', 'Frutta', 'Acqua', '', '', 'glutine',
       '6,50', '30', 'SI', 'NO', '520', '16', '70', '15']]),
]
MODELLI_MAP = {m[0]: m for m in MODELLI}


# ── Utilità ─────────────────────────────────────────────────────────────────

def _norm(s):
    """Intestazione normalizzata: minuscola, senza accenti, spazi e asterischi."""
    s = unicodedata.normalize('NFKD', str(s or ''))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'[^a-z0-9]+', ' ', s.lower().replace('*', '')).strip()


def _numero(v, intero=False):
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return int(round(v)) if intero else float(v)
    s = str(v).strip().replace('€', '').replace(' ', '').replace(',', '.')
    if not s:
        return None
    f = float(s)
    return int(round(f)) if intero else f


def _sino(v, default=False):
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s == '':
        return default
    if s in SI:
        return True
    if s in NO:
        return False
    raise ValueError('valore "%s" non riconosciuto: usa SI o NO' % v)


def _allergeni(v):
    if not v:
        return ''
    etichette = {_norm(l): k for k, l, _i in ALLERGENS}
    chiavi = {k for k, _l, _i in ALLERGENS}
    out = []
    for pezzo in str(v).replace(';', ',').split(','):
        p = pezzo.strip()
        if not p:
            continue
        k = p.lower().replace(' ', '_')
        if k in chiavi:
            out.append(k)
        elif _norm(p) in etichette:
            out.append(etichette[_norm(p)])
        else:
            raise ValueError('allergene "%s" sconosciuto (ammessi: %s)' % (p, ', '.join(sorted(chiavi))))
    return ','.join(dict.fromkeys(out))


def _data(v):
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d.%m.%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError('data "%s" non valida: usa GG/MM/AAAA' % s)


def _testo(v, massimo=None):
    if v is None:
        return ''
    s = str(v).strip()
    if isinstance(v, float) and v.is_integer():
        s = str(int(v))
    return s[:massimo] if massimo else s


# ── Modello .xlsx ───────────────────────────────────────────────────────────

def genera_modello(tipo, tenant_id=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    chiave, etichetta, _icona, descrizione, colonne, esempi = MODELLI_MAP[tipo]
    wb = Workbook()
    ws = wb.active
    ws.title = 'Dati'
    testa_font = Font(bold=True, color='FFFFFF')
    testa_fill = PatternFill('solid', fgColor='0F3460')
    obbl_fill = PatternFill('solid', fgColor='E94560')
    for i, (nome, obbl, _sp, _tipo) in enumerate(colonne, 1):
        c = ws.cell(row=1, column=i, value=nome + (' *' if obbl else ''))
        c.font = testa_font
        c.fill = obbl_fill if obbl else testa_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(i)].width = max(14, min(40, len(nome) + 6))
    ws.freeze_panes = 'A2'
    # Convalide: SI/NO e gli elenchi chiusi
    for i, (nome, _o, _sp, t) in enumerate(colonne, 1):
        col = get_column_letter(i)
        if t == 'sino':
            dv = DataValidation(type='list', formula1='"SI,NO"', allow_blank=True)
        elif t.startswith('elenco:'):
            dv = DataValidation(type='list', formula1='"%s"' % t.split(':', 1)[1], allow_blank=True)
        else:
            continue
        dv.error = 'Scegli un valore dall\'elenco'
        dv.errorTitle = 'Valore non ammesso'
        ws.add_data_validation(dv)
        dv.add('%s2:%s500' % (col, col))

    es = wb.create_sheet('Esempio')
    for i, (nome, obbl, _sp, _t) in enumerate(colonne, 1):
        c = es.cell(row=1, column=i, value=nome + (' *' if obbl else ''))
        c.font = testa_font
        c.fill = obbl_fill if obbl else testa_fill
        es.column_dimensions[get_column_letter(i)].width = max(14, min(40, len(nome) + 6))
    for r, riga in enumerate(esempi, 2):
        for i, v in enumerate(riga, 1):
            es.cell(row=r, column=i, value=v)
    es.cell(row=len(esempi) + 3, column=1,
            value='Queste righe sono solo un esempio: compila il foglio "Dati".').font = Font(italic=True, color='888888')

    ist = wb.create_sheet('Istruzioni')
    ist.column_dimensions['A'].width = 26
    ist.column_dimensions['B'].width = 14
    ist.column_dimensions['C'].width = 90
    ist['A1'] = 'QuickLunch — modello: ' + etichetta
    ist['A1'].font = Font(bold=True, size=14, color='0F3460')
    ist['A2'] = descrizione
    ist['A2'].alignment = Alignment(wrap_text=True)
    ist.merge_cells('A2:C2')
    ist.row_dimensions[2].height = 40
    ist['A4'] = 'Come si usa'
    ist['A4'].font = Font(bold=True)
    passi = ['1. Compila il foglio "Dati": una riga per elemento, l\'intestazione non si tocca.',
             '2. Le colonne con l\'asterisco (in rosso) sono obbligatorie; le altre puoi lasciarle vuote o toglierle.',
             '3. Una riga con lo stesso nome di un elemento già presente lo aggiorna, non lo duplica.',
             '4. Salva in formato .xlsx e caricalo da Prodotti → Importa da Excel. Con "Solo verifica" controlli senza scrivere.',
             '5. Le righe con errori vengono saltate e spiegate nel resoconto: correggile e ricarica solo quelle.']
    for r, p in enumerate(passi, 5):
        ist.cell(row=r, column=1, value=p)
        ist.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r0 = 5 + len(passi) + 1
    for i, t in enumerate(('Colonna', 'Obbligatoria', 'Che cosa scrivere'), 1):
        c = ist.cell(row=r0, column=i, value=t)
        c.font = testa_font
        c.fill = testa_fill
    for r, (nome, obbl, spieg, t) in enumerate(colonne, r0 + 1):
        ist.cell(row=r, column=1, value=nome)
        ist.cell(row=r, column=2, value='sì' if obbl else 'no')
        extra = ''
        if t == 'sino':
            extra = ' Valori: SI / NO.'
        elif t.startswith('elenco:'):
            extra = ' Valori: %s.' % t.split(':', 1)[1].replace(',', ', ')
        elif t == 'numero':
            extra = ' Numero, con virgola o punto.'
        elif t == 'intero':
            extra = ' Numero intero.'
        ist.cell(row=r, column=3, value=(spieg + extra).strip()).alignment = Alignment(wrap_text=True)
    r = r0 + len(colonne) + 2
    ist.cell(row=r, column=1, value='Allergeni ammessi').font = Font(bold=True)
    ist.cell(row=r + 1, column=1, value=', '.join('%s (%s)' % (k, l) for k, l, _i in ALLERGENS))
    ist.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=3)
    ist.row_dimensions[r + 1].height = 45
    ist.cell(row=r + 1, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    r += 3
    esistenti = _elenchi_esistenti(tipo, tenant_id)
    for titolo, voci in esistenti:
        ist.cell(row=r, column=1, value=titolo).font = Font(bold=True)
        ist.cell(row=r + 1, column=1, value=', '.join(voci) if voci else '(nessuno ancora)')
        ist.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=3)
        ist.cell(row=r + 1, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ist.row_dimensions[r + 1].height = 30
        r += 3

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _elenchi_esistenti(tipo, tenant_id):
    """Le voci già presenti che aiutano a compilare (categorie, aziende...)."""
    filtro = {'tenant_id': tenant_id} if tenant_id is not None else {}
    if tipo == 'prodotti':
        return [('Categorie già presenti', sorted({c.name for c in Category.query.filter_by(**filtro).all()}))]
    if tipo == 'ingredienti':
        return [('Categorie già presenti (builder)',
                 sorted({'%s (%s)' % (c.name, c.builder_type)
                         for c in IngredientCategory.query.filter_by(**filtro).all()}))]
    if tipo == 'consumabili':
        return [('Fornitori già presenti', sorted({s.name for s in Supplier.query.filter_by(**filtro).all()}))]
    if tipo in ('clienti', 'pasti_convenzione'):
        return [('Convenzioni aziendali esistenti',
                 sorted({c.name for c in CorporateAccount.query.filter_by(**filtro).all()}))]
    return []


# ── Importazione ────────────────────────────────────────────────────────────

def _leggi_righe(contenuto, colonne):
    """[(numero_riga, {chiave_colonna: valore})] dal foglio Dati (o dal primo)."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(contenuto), data_only=True, read_only=True)
    ws = wb['Dati'] if 'Dati' in wb.sheetnames else wb.worksheets[0]
    righe = list(ws.iter_rows(values_only=True))
    if not righe:
        raise ValueError('Il foglio è vuoto.')
    attese = {_norm(nome): nome for nome, _o, _s, _t in colonne}
    posizioni = {}
    for i, cella in enumerate(righe[0]):
        n = _norm(cella)
        if n in attese:
            posizioni[attese[n]] = i
    mancanti = [nome for nome, obbl, _s, _t in colonne if obbl and nome not in posizioni]
    if mancanti:
        raise ValueError('Mancano le colonne obbligatorie: %s. Usa il modello scaricato da qui.'
                         % ', '.join(mancanti))
    out = []
    for n, riga in enumerate(righe[1:], 2):
        if riga is None or all(v is None or str(v).strip() == '' for v in riga):
            continue
        d = {}
        for nome, i in posizioni.items():
            d[nome] = riga[i] if i < len(riga) else None
        out.append((n, d))
    return out


def importa(tipo, contenuto, tenant_id, solo_verifica=False):
    """Importa (o verifica soltanto). Ritorna il resoconto."""
    chiave, etichetta, _ic, _d, colonne, _es = MODELLI_MAP[tipo]
    esito = {'tipo': tipo, 'etichetta': etichetta, 'creati': 0, 'aggiornati': 0,
             'errori': [], 'avvisi': [], 'righe': 0, 'password': [], 'solo_verifica': solo_verifica}
    try:
        righe = _leggi_righe(contenuto, colonne)
    except Exception as exc:                       # file non Excel, foglio vuoto, colonne mancanti
        esito['errori'].append((0, str(exc) or 'File non leggibile: serve un .xlsx'))
        return esito
    esito['righe'] = len(righe)
    gestore = {'prodotti': _importa_prodotto, 'ingredienti': _importa_ingrediente,
               'banco': _importa_banco, 'consumabili': _importa_consumabile,
               'clienti': _importa_cliente, 'pasti_convenzione': _importa_pasto}[tipo]
    # Una transazione per riga: la riga buona si scrive subito, quella con
    # errori si annulla da sola senza toccare le altre. Niente SAVEPOINT,
    # che con SQLite (pysqlite) non e' affidabile. In sola verifica non si
    # fa nessun commit e alla fine si annulla tutto.
    for n, d in righe:
        for nome, obbl, _s, _t in colonne:
            if obbl and _testo(d.get(nome)) == '':
                esito['errori'].append((n, 'manca "%s"' % nome))
                break
        else:
            try:
                creato = gestore(d, tenant_id, esito)
                if solo_verifica:
                    db.session.flush()
                else:
                    db.session.commit()
                esito['creati' if creato else 'aggiornati'] += 1
            except Exception as exc:
                db.session.rollback()
                esito['errori'].append((n, str(exc) or exc.__class__.__name__))
    if solo_verifica:
        db.session.rollback()
        esito['password'] = []
    return esito


def _nutrizione(obj, d):
    obj.kcal = _numero(d.get('Kcal'), intero=True)
    obj.proteine_g = _numero(d.get('Proteine g'))
    obj.carboidrati_g = _numero(d.get('Carboidrati g'))
    obj.grassi_g = _numero(d.get('Grassi g'))
    if 'Vegetariano' in d:
        obj.is_vegetarian = _sino(d.get('Vegetariano'), getattr(obj, 'is_vegetarian', False) or False)
    if 'Vegano' in d:
        obj.is_vegan = _sino(d.get('Vegano'), getattr(obj, 'is_vegan', False) or False)
        if obj.is_vegan:
            obj.is_vegetarian = True


def _importa_prodotto(d, tid, esito):
    nome = _testo(d['Nome'], 128)
    nome_cat = _testo(d['Categoria'], 64)
    cat = Category.query.filter(Category.tenant_id == tid, db.func.lower(Category.name) == nome_cat.lower()).first()
    if cat is None:
        cat = Category(name=nome_cat, tenant_id=tid)
        db.session.add(cat)
        db.session.flush()
        esito['avvisi'].append('creata la categoria "%s"' % nome_cat)
    prezzo = _numero(d['Prezzo'])
    if prezzo is None or prezzo < 0:
        raise ValueError('prezzo non valido')
    p = Product.query.filter(Product.tenant_id == tid, db.func.lower(Product.name) == nome.lower()).first()
    creato = p is None
    if creato:
        p = Product(name=nome, tenant_id=tid, price=prezzo, category_id=cat.id)
        db.session.add(p)
    p.price = prezzo
    p.category_id = cat.id
    if 'Descrizione' in d:
        p.description = _testo(d.get('Descrizione'))
    if 'Allergeni' in d:
        p.allergens = _allergeni(d.get('Allergeni'))
    q = _numero(d.get('Quantita giornaliera'), intero=True)
    if q is not None:
        p.daily_quantity = max(0, q)
    if 'Codice a barre' in d:
        p.barcode = _testo(d.get('Codice a barre'), 32) or None
    p.is_active = _sino(d.get('Attivo'), True)
    _nutrizione(p, d)
    return creato


def _importa_ingrediente(d, tid, esito):
    nome = _testo(d['Nome'], 128)
    builder = _testo(d['Builder']).lower()
    if builder not in ('panino', 'insalata', 'poke'):
        raise ValueError('builder "%s" non valido: panino, insalata o poke' % builder)
    nome_cat = _testo(d['Categoria'], 64)
    cat = IngredientCategory.query.filter(
        IngredientCategory.tenant_id == tid, IngredientCategory.builder_type == builder,
        db.func.lower(IngredientCategory.name) == nome_cat.lower()).first()
    if cat is None:
        cat = IngredientCategory(name=nome_cat, builder_type=builder, tenant_id=tid)
        db.session.add(cat)
        db.session.flush()
        esito['avvisi'].append('creata la categoria "%s" (%s)' % (nome_cat, builder))
    ing = Ingredient.query.filter(Ingredient.tenant_id == tid, Ingredient.category_id == cat.id,
                                  db.func.lower(Ingredient.name) == nome.lower()).first()
    creato = ing is None
    if creato:
        ing = Ingredient(name=nome, tenant_id=tid, category_id=cat.id)
        db.session.add(ing)
    ing.price_extra = _numero(d.get('Prezzo extra')) or 0.0
    if 'Allergeni' in d:
        ing.allergens = _allergeni(d.get('Allergeni'))[:128]
    g = _numero(d.get('Grammi per porzione'))
    if g is not None:
        ing.grams_per_serving = g
    ing.is_active = _sino(d.get('Attivo'), True)
    _nutrizione(ing, d)
    return creato


def _importa_banco(d, tid, esito):
    nome = _testo(d['Nome'], 64)
    prezzo = _numero(d['Prezzo'])
    if prezzo is None or prezzo < 0:
        raise ValueError('prezzo non valido')
    b = BancoItem.query.filter(BancoItem.tenant_id == tid, db.func.lower(BancoItem.name) == nome.lower()).first()
    creato = b is None
    if creato:
        b = BancoItem(name=nome, tenant_id=tid, price=prezzo)
        db.session.add(b)
    b.price = prezzo
    icona = _testo(d.get('Icona'), 64)
    if icona:
        b.icon = icona if icona.startswith('fa-') else 'fa-' + icona
    colore = _testo(d.get('Colore')).lower()
    if colore:
        if colore not in ('primary', 'secondary', 'success', 'danger', 'warning', 'info', 'dark'):
            raise ValueError('colore "%s" non valido' % colore)
        b.color = colore
    o = _numero(d.get('Ordine'), intero=True)
    if o is not None:
        b.sort_order = o
    b.is_active = _sino(d.get('Attivo'), True)
    return creato


def _importa_consumabile(d, tid, esito):
    nome = _testo(d['Nome'], 128)
    c = ConsumableItem.query.filter(ConsumableItem.tenant_id == tid,
                                    db.func.lower(ConsumableItem.name) == nome.lower()).first()
    creato = c is None
    if creato:
        c = ConsumableItem(name=nome, tenant_id=tid)
        db.session.add(c)
    unita = _testo(d.get('Unita'), 20)
    if unita:
        c.unit = unita
    q = _numero(d.get('Quantita'))
    if q is not None:
        c.quantity = q
    s = _numero(d.get('Soglia minima'))
    if s is not None:
        c.min_threshold = s
    nome_f = _testo(d.get('Fornitore'), 128)
    if nome_f:
        f = Supplier.query.filter(Supplier.tenant_id == tid, db.func.lower(Supplier.name) == nome_f.lower()).first()
        if f is None:
            f = Supplier(name=nome_f, tenant_id=tid)
            db.session.add(f)
            db.session.flush()
            esito['avvisi'].append('creato il fornitore "%s"' % nome_f)
        c.supplier_id = f.id
    c.alert_active = bool(c.min_threshold and c.quantity <= c.min_threshold)
    return creato


def _importa_cliente(d, tid, esito):
    email = _testo(d['Email'], 120).lower()
    if '@' not in email or '.' not in email.split('@')[-1]:
        raise ValueError('email "%s" non valida' % email)
    u = User.query.filter(db.func.lower(User.email) == email).first()
    creato = u is None
    if creato:
        username = email
        if User.query.filter_by(username=username).first():
            username = email.split('@')[0] + '_' + secrets.token_hex(2)
        u = User(username=username, email=email, is_client=True, tenant_id=tid)
        pw = secrets.token_urlsafe(8)
        u.set_password(pw)
        esito['password'].append((email, pw))
        db.session.add(u)
    elif u.tenant_id not in (None, tid):
        raise ValueError('l\'email appartiene a un altro locale')
    else:
        u.is_client = True
    u.first_name = _testo(d['Nome'], 64)
    u.last_name = _testo(d['Cognome'], 64)
    if 'Telefono' in d:
        u.phone = _testo(d.get('Telefono'), 30)
    if 'Reparto' in d:
        u.reparto = _testo(d.get('Reparto'), 120)
    if 'Data di nascita' in d:
        dn = _data(d.get('Data di nascita'))
        if dn is not None:
            u.birth_date = dn
    u.is_active = _sino(d.get('Attivo'), True)
    azienda = _testo(d.get('Azienda convenzionata'), 128)
    if azienda:
        corp = CorporateAccount.query.filter(
            CorporateAccount.tenant_id == tid,
            db.func.lower(CorporateAccount.name) == azienda.lower()).first()
        if corp is None:
            raise ValueError('convenzione "%s" non trovata: creala prima da Convenzioni' % azienda)
        db.session.flush()
        m = CorporateMembership.query.filter_by(user_id=u.id).first()
        if m is None:
            db.session.add(CorporateMembership(user_id=u.id, corporate_id=corp.id, is_active=True))
        else:
            m.corporate_id = corp.id
            m.is_active = True
    return creato


def _importa_pasto(d, tid, esito):
    azienda = _testo(d['Azienda'], 128)
    corp = CorporateAccount.query.filter(
        CorporateAccount.tenant_id == tid,
        db.func.lower(CorporateAccount.name) == azienda.lower()).first()
    if corp is None:
        raise ValueError('convenzione "%s" non trovata: creala prima da Convenzioni' % azienda)
    nome = _testo(d['Nome'], 128)
    mc = MealConfiguration.query.filter(
        MealConfiguration.corporate_id == corp.id,
        db.func.lower(MealConfiguration.name) == nome.lower()).first()
    creato = mc is None
    if creato:
        mc = MealConfiguration(corporate_id=corp.id, name=nome, tenant_id=tid)
        db.session.add(mc)
    for campo, col in (('primo', 'Primo'), ('secondo', 'Secondo'), ('contorno', 'Contorno'),
                       ('bevanda', 'Bevanda'), ('caffe', 'Caffe'), ('description', 'Descrizione')):
        if col in d:
            setattr(mc, campo, _testo(d.get(col), 256 if campo != 'description' else None))
    if 'Allergeni' in d:
        mc.allergens = _allergeni(d.get('Allergeni'))
    mc.price = _numero(d.get('Prezzo'))
    mc.max_bookings = _numero(d.get('Max prenotazioni'), intero=True)
    _nutrizione(mc, d)
    return creato
